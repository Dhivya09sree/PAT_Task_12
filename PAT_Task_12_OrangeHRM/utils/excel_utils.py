from openpyxl import load_workbook

class ExcelHandler:
    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.active
#gdet the data from exel sheet
    def get_test_data(self):
        test_data = []
        for row in range(2, self.sheet.max_row + 1):
            data = {
                "test_id": self.sheet.cell(row, 1).value,
                "username": self.sheet.cell(row, 2).value,
                "password": self.sheet.cell(row, 3).value,
                "date": self.sheet.cell(row, 4).value,
                "time_of_test": self.sheet.cell(row, 5).value,
                "tester_name": self.sheet.cell(row, 6).value,
            }
            test_data.append(data)
        return test_data
  #write the data in ecel sheet
    def write_test_result(self, row, result):
        self.sheet.cell(row + 2, 7).value = result
        self.workbook.save(self.file_path)
